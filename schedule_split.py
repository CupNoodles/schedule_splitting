import openpyxl
import json 
import re
import numpy


max_ratio = 0.55
min_ratio = 0.45
max_class = 18

#read classes from schedule
mastersched = openpyxl.load_workbook(filename = "ms.xlsx")
ms = mastersched["MasterScheduleFinal 2019-2020"]

classes = {}

for row in ms.values:
    if row[0] != 'Course Code' and row[3] != None and row[7] > 0:
        classes[str(row[0]) + "/" + str(row[1])] = {'student_count' : row[7], 'students' : [], 'A_student_count' : 0, 'B_student_count': 0}

#create random a/b students list

sr = openpyxl.load_workbook(filename = "ls.xlsx")["LineSchedule 2019-2020"]



s_count = len(tuple(sr.rows))
s_map = {}
s_sched = {}

i = 0
for row in sr.values:
    i = i+1
    if row[1] != 'StudentId':

        if(i < s_count / 2):
            s_map[row[1]] = 'A'
        else:
            s_map[row[1]] = 'B'
        
        s_sched[row[1]] = []
        for j in range(5, 15):
            if row[j] != None:
                text = row[j].split("\n")
                for k in range(0, len(text)):
                    class_found = re.search('^[0-9A-Z]+/[0-9]+$', text[k])
                    if class_found:
                        c = class_found.group(0)
                        
                        s_sched[row[1]].append(c)
                        if c in classes.keys():
                            classes[c]['students'].append(row[1])

for student in s_sched.keys():
    s_sched[student] = set(s_sched[student])

for c in classes.keys():
    classes[c]['students'] = set(classes[c]['students'])

#measure class size good/bad ratio
def a_b_ratio():
    good = 0
    bad = 0

    for k in classes:
        classes[k]['A_student_count'] = 0
        classes[k]['B_student_count'] = 0
        for j in range(0, len(classes[k]['students'])):
            if s_map[classes[k]['students'][j]] == 'A':
                classes[k]['A_student_count'] += 1
            else:
                classes[k]['B_student_count'] += 1

    for k in classes:
        
        if classes[k]['student_count'] > 0 :
            ratio = float(classes[k]['A_student_count']) / classes[k]['student_count'] 

            if (ratio < max_ratio and ratio > min_ratio) or (classes[k]['A_student_count'] < max_class and classes[k]['B_student_count'] < max_class) :
                good += 1
            else :
                bad += 1
    
    return good / len(classes)

def class_distance(k):

    cd = 0

    if classes[k]['student_count'] > (2 * max_class):
        mc = float(classes[k]['student_count']) * max_ratio
    else:
        mc = max_class
    
    if classes[k]['A_student_count'] > mc:
        cd += classes[k]['A_student_count'] - mc
    if classes[k]['B_student_count'] > mc:
        cd += classes[k]['B_student_count'] - mc
    
    return cd

#same algo, measure distance instead of bool
def a_b_distance():
    badness = 0

    for k in classes:
        classes[k]['A_student_count'] = 0
        classes[k]['B_student_count'] = 0
        for j in classes[k]['students']:
            if s_map[j] == 'A':
                classes[k]['A_student_count'] += 1
            else:
                classes[k]['B_student_count'] += 1

    for k in classes:
        badness += class_distance(k)
    
    return badness



def switch_kids(a, b):
    if s_map[a] == 'A':
        s_map[a] = 'B'
    else:
        s_map[a] = 'A'

    if s_map[b] == 'A':
        s_map[b] = 'B'
    else:
        s_map[b] = 'A'
        


badness = a_b_distance()
print(badness)

while(badness > 1):

    print(list(classes.keys()))
    keys = numpy.random.permutation(list(classes.keys()))

    #find an A student in a class with too many
    for c in keys:
        cd = class_distance(c)
        print(" begin processing class " + c + " class distance = " + str(cd))
        if cd > 0:
            a_students = []
            b_students = []
            for student in classes[c]['students']:
                if s_map[student] == 'A':
                    a_students.append(student)
                if s_map[student] == 'B':
                    b_students.append(student)
            
            a_students = numpy.random.permutation(a_students)
            b_students = numpy.random.permutation(b_students)

            for a in a_students:
                for b in b_students:
                    old_badness = badness
                    print(c + " switching student " + str(a) + " and student " + str(b) + " class distance is " + str(class_distance(c)))
                    switch_kids(a, b)
                    badness = a_b_distance()
                    print("distance delta " + str( int(badness) - int(old_badness)))
                    if badness <= old_badness:
                        print("badness delta OK. keeping change.")
                        print("total badness: " + str(badness) )
                        if class_distance(c) == 0:
                            break
                    else:
                        print("badness delta fail. switching back " + str(a) + " and student " + str(b))
                        print("total badness: " + str(old_badness) + " (old)")
                        switch_kids(a, b)
                        badness = old_badness                        
                        


                    
                if class_distance(c) == 0 :
                    break

dest_filename = 'student_mappings.xlsx'
output = openpyxl.Workbook()

os = output.active
os.title = "Student Mappings"
os.append(['Student ID', 'A/B'])
for student in s_map.keys():
    os.append([student, s_map[student]])
output.save(filename = dest_filename)

a_b_distance()

for c in classes:
    if classes[c]["student_count"] > (max_class * 2):
        print (c + " A: " + str(classes[c]["A_student_count"]) + " B: " + str(classes[c]["B_student_count"]) + " ratio = " + str((float(classes[c]["A_student_count"]) / classes[c]["student_count"]) * 100) + "%")